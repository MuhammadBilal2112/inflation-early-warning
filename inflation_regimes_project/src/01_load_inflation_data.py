"""
==============================================================================
STEP 2: LOAD AND CLEAN THE INFLATION DATA
==============================================================================

Reads all inflation data from the World Bank Excel file, handles the
different column structures across sheets, computes inflation rates from
monthly price indices, and produces two clean CSV files.

OUTPUT FILES:
  - data/processed/inflation_panel_monthly.csv
  - data/processed/inflation_panel_annual.csv

RUN: python3 Step02_Load_Inflation_Data.py
TIME: 2-3 minutes
==============================================================================
"""

import pandas as pd
import numpy as np
import openpyxl
import os
import warnings
warnings.filterwarnings('ignore')


# ============================================================
# Paths
# ============================================================

BASE_DIR = "inflation_regimes_project"

RAW_DIR = os.path.join(BASE_DIR, "data", "raw")
PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
INTERIM_DIR = os.path.join(BASE_DIR, "data", "interim")

inflation_file = os.path.join(RAW_DIR, "Inflationdata.xlsx")
if os.path.exists(inflation_file):
    size_mb = os.path.getsize(inflation_file) / (1024*1024)
    print(f"Found inflation data file: {size_mb:.1f} MB")
else:
    print(f"ERROR: Cannot find {inflation_file}")
    print(f"Current directory: {os.getcwd()}")


# ============================================================
# Annual data loader
# ============================================================
# Annual sheets contain inflation RATES directly (not indices).
# Structure: Country Code | IMF Code | Country | Indicator | Series | 1970 | 1971 | ...

def load_annual_sheet(filepath, sheet_name, measure_name):
    """Load one annual inflation sheet and reshape from wide to long format."""
    print(f"  Loading {sheet_name} ({measure_name})...", end="", flush=True)

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[sheet_name]

    header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    year_info = []
    for i, h in enumerate(header):
        if isinstance(h, (int, float)) and 1960 <= h <= 2030:
            year_info.append((i, int(h)))

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        country_code = row[0]
        country_name = row[2]

        if not country_code or not country_name:
            continue

        for col_idx, year in year_info:
            value = row[col_idx] if col_idx < len(row) else None

            if value is not None and isinstance(value, (int, float)):
                records.append({
                    'country_code': str(country_code).strip(),
                    'country_name': str(country_name).strip(),
                    'year': year,
                    measure_name: float(value)
                })

    wb.close()

    df = pd.DataFrame(records)
    n_countries = df['country_code'].nunique()
    n_years = df['year'].nunique()
    print(f" {n_countries} countries, {n_years} years, {len(df):,} observations")

    return df


# ============================================================
# Load annual sheets
# ============================================================

print("=" * 60)
print("LOADING ANNUAL INFLATION DATA")
print("=" * 60)

annual_sheets = {
    'hcpi_a': 'hcpi',
    'fcpi_a': 'fcpi',
    'ecpi_a': 'ecpi',
    'ccpi_a': 'ccpi',
    'ppi_a': 'ppi',
    'def_a': 'gdp_deflator',
}

annual_dfs = {}
for sheet_name, measure_name in annual_sheets.items():
    annual_dfs[measure_name] = load_annual_sheet(inflation_file, sheet_name, measure_name)

print("\nMerging all annual measures...")

annual_panel = annual_dfs['hcpi'].copy()

for measure_name, df in annual_dfs.items():
    if measure_name == 'hcpi':
        continue
    annual_panel = annual_panel.merge(
        df[['country_code', 'year', measure_name]],
        on=['country_code', 'year'],
        how='outer'
    )

name_map = {}
for df in annual_dfs.values():
    for _, row in df[['country_code', 'country_name']].drop_duplicates().iterrows():
        name_map[row['country_code']] = row['country_name']

annual_panel['country_name'] = annual_panel['country_code'].map(name_map)
annual_panel = annual_panel.sort_values(['country_code', 'year']).reset_index(drop=True)

print(f"\nAnnual panel created:")
print(f"  Shape: {annual_panel.shape}")
print(f"  Countries: {annual_panel['country_code'].nunique()}")
print(f"  Years: {annual_panel['year'].min()} to {annual_panel['year'].max()}")
print(f"\n  Missing values per measure:")
for col in ['hcpi', 'fcpi', 'ecpi', 'ccpi', 'ppi', 'gdp_deflator']:
    n_total = len(annual_panel)
    n_missing = annual_panel[col].isna().sum()
    pct = n_missing / n_total * 100
    print(f"    {col:15s}: {n_missing:,}/{n_total:,} missing ({pct:.1f}%)")


# ============================================================
# Monthly data loader
# ============================================================
# Monthly sheets contain price INDEX values, not rates.
# YoY inflation is computed as: (Index_t / Index_{t-12} - 1) * 100
#
# Note: fcpi_m has 6 metadata columns vs 5 for all other sheets.
# This function detects the time column position automatically.

def load_monthly_sheet(filepath, sheet_name, measure_name):
    """Load one monthly price index sheet and compute YoY inflation rates."""
    print(f"  Loading {sheet_name} ({measure_name})...", end="", flush=True)

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[sheet_name]

    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    # Time columns are integers > 190000 (format YYYYMM)
    time_start_idx = None
    for i, h in enumerate(header):
        if isinstance(h, (int, float)) and h > 190000:
            time_start_idx = i
            break

    if time_start_idx is None:
        print(" ERROR: Could not find time columns!")
        wb.close()
        return pd.DataFrame()

    time_columns = []
    for i in range(time_start_idx, len(header)):
        h = header[i]
        if isinstance(h, (int, float)) and h > 190000:
            time_columns.append((i, int(h)))

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        country_code = row[0]
        country_name = row[2]

        if not country_code or not country_name:
            continue

        cc = str(country_code).strip()
        cn = str(country_name).strip()

        for col_idx, date_code in time_columns:
            value = row[col_idx] if col_idx < len(row) else None

            if value is not None and isinstance(value, (int, float)):
                records.append({
                    'country_code': cc,
                    'country_name': cn,
                    'date_code': date_code,
                    f'{measure_name}_index': float(value)
                })

    wb.close()

    df = pd.DataFrame(records)

    if df.empty:
        print(" No data found!")
        return df

    # Convert date_code (YYYYMM) to proper date
    df['year'] = df['date_code'] // 100
    df['month'] = df['date_code'] % 100
    df['date'] = pd.to_datetime(
        df['year'].astype(str) + '-' + df['month'].astype(str).str.zfill(2) + '-01'
    )

    df = df.sort_values(['country_code', 'date']).reset_index(drop=True)

    # Compute YoY: (Index_t / Index_{t-12} - 1) * 100
    df[f'{measure_name}_index_lag12'] = df.groupby('country_code')[f'{measure_name}_index'].shift(12)
    df[f'{measure_name}_yoy'] = (
        (df[f'{measure_name}_index'] / df[f'{measure_name}_index_lag12']) - 1
    ) * 100

    df = df.drop(columns=[f'{measure_name}_index_lag12', 'date_code', 'year', 'month'])

    n_countries = df['country_code'].nunique()
    n_months = df['date'].nunique()
    n_yoy = df[f'{measure_name}_yoy'].notna().sum()
    print(f" {n_countries} countries, {n_months} months, {n_yoy:,} YoY rates computed")

    return df


# ============================================================
# Load monthly sheets
# ============================================================

print("\n" + "=" * 60)
print("LOADING MONTHLY INFLATION DATA")
print("=" * 60)

monthly_sheets = {
    'hcpi_m': 'hcpi',
    'fcpi_m': 'fcpi',
    'ecpi_m': 'ecpi',
    'ccpi_m': 'ccpi',
    'ppi_m': 'ppi',
}

monthly_dfs = {}
for sheet_name, measure_name in monthly_sheets.items():
    monthly_dfs[measure_name] = load_monthly_sheet(inflation_file, sheet_name, measure_name)


# ============================================================
# Merge monthly measures
# ============================================================

print("\nMerging all monthly measures...")

monthly_panel = monthly_dfs['hcpi'][['country_code', 'country_name', 'date',
                                      'hcpi_index', 'hcpi_yoy']].copy()

for measure_name, df in monthly_dfs.items():
    if measure_name == 'hcpi':
        continue

    cols_to_merge = ['country_code', 'date', f'{measure_name}_index', f'{measure_name}_yoy']
    cols_to_merge = [c for c in cols_to_merge if c in df.columns]

    monthly_panel = monthly_panel.merge(
        df[cols_to_merge],
        on=['country_code', 'date'],
        how='outer'
    )

name_map_monthly = {}
for df in monthly_dfs.values():
    if 'country_code' in df.columns and 'country_name' in df.columns:
        for _, row in df[['country_code', 'country_name']].drop_duplicates().iterrows():
            name_map_monthly[row['country_code']] = row['country_name']

monthly_panel['country_name'] = monthly_panel['country_code'].map(name_map_monthly)
monthly_panel = monthly_panel.sort_values(['country_code', 'date']).reset_index(drop=True)

print(f"\nMonthly panel created:")
print(f"  Shape: {monthly_panel.shape}")
print(f"  Countries: {monthly_panel['country_code'].nunique()}")
print(f"  Date range: {monthly_panel['date'].min().strftime('%Y-%m')} to {monthly_panel['date'].max().strftime('%Y-%m')}")
print(f"\n  Missing values per YoY measure:")
for col in ['hcpi_yoy', 'fcpi_yoy', 'ecpi_yoy', 'ccpi_yoy', 'ppi_yoy']:
    if col in monthly_panel.columns:
        n_total = len(monthly_panel)
        n_missing = monthly_panel[col].isna().sum()
        pct = n_missing / n_total * 100
        print(f"    {col:15s}: {n_missing:,}/{n_total:,} missing ({pct:.1f}%)")


# ============================================================
# Validation
# ============================================================

print("\n" + "=" * 60)
print("VALIDATION: Checking computed values against known facts")
print("=" * 60)

validation_passed = True

print("\n--- Test 1: UK Headline CPI ---")
uk_monthly = monthly_panel[monthly_panel['country_code'] == 'GBR'].copy()
uk_monthly = uk_monthly.set_index('date')

for year in [2019, 2020, 2021, 2022, 2023]:
    year_data = uk_monthly[uk_monthly.index.year == year]['hcpi_yoy']
    if not year_data.empty:
        avg = year_data.mean()
        print(f"  UK {year} average YoY inflation: {avg:.2f}%", end="")

        expected = {2019: (1.0, 2.5), 2020: (0.5, 1.5), 2021: (1.5, 4.0),
                    2022: (5.0, 11.0), 2023: (5.0, 10.0)}
        lo, hi = expected[year]
        if lo <= avg <= hi:
            print("  [OK]")
        else:
            print(f"  [WARNING: expected {lo}-{hi}%]")
            validation_passed = False
    else:
        print(f"  UK {year}: no data")

print("\n--- Test 2: US Headline CPI ---")
us_monthly = monthly_panel[monthly_panel['country_code'] == 'USA'].copy()
us_monthly = us_monthly.set_index('date')

for year in [2020, 2021, 2022]:
    year_data = us_monthly[us_monthly.index.year == year]['hcpi_yoy']
    if not year_data.empty:
        avg = year_data.mean()
        print(f"  US {year} average YoY inflation: {avg:.2f}%", end="")
        expected_us = {2020: (0.5, 2.0), 2021: (3.0, 6.0), 2022: (6.0, 9.0)}
        lo, hi = expected_us[year]
        if lo <= avg <= hi:
            print("  [OK]")
        else:
            print(f"  [WARNING: expected {lo}-{hi}%]")
            validation_passed = False

print("\n--- Test 3: Turkey (high inflation check) ---")
tr_monthly = monthly_panel[monthly_panel['country_code'] == 'TUR'].copy()
tr_monthly = tr_monthly.set_index('date')

for year in [2021, 2022]:
    year_data = tr_monthly[tr_monthly.index.year == year]['hcpi_yoy']
    if not year_data.empty:
        avg = year_data.mean()
        print(f"  Turkey {year} average YoY inflation: {avg:.2f}%", end="")
        expected_tr = {2021: (15, 25), 2022: (50, 85)}
        lo, hi = expected_tr[year]
        if lo <= avg <= hi:
            print("  [OK]")
        else:
            print(f"  [WARNING: expected {lo}-{hi}%]")
            validation_passed = False

if validation_passed:
    print("\n  All validation checks passed.")
else:
    print("\n  Some checks raised warnings — review the values above.")
    print("  Small discrepancies are normal (rounding, index base differences).")


# ============================================================
# Data quality checks
# ============================================================

print("\n" + "=" * 60)
print("DATA QUALITY CHECKS")
print("=" * 60)

for col in ['hcpi_yoy', 'fcpi_yoy', 'ecpi_yoy']:
    if col not in monthly_panel.columns:
        continue
    data = monthly_panel[col].dropna()
    extreme_high = (data > 100).sum()
    extreme_low = (data < -50).sum()
    print(f"\n  {col}:")
    print(f"    Range: {data.min():.1f}% to {data.max():.1f}%")
    print(f"    Observations > 100% YoY: {extreme_high}")
    print(f"    Observations < -50% YoY: {extreme_low}")

    if extreme_high > 0:
        extreme_rows = monthly_panel[monthly_panel[col] > 100][['country_code', 'country_name', 'date', col]]
        top_countries = extreme_rows.groupby('country_name').size().sort_values(ascending=False).head(5)
        print(f"    Top countries with >100% inflation:")
        for country, count in top_countries.items():
            print(f"      {country}: {count} months")

monthly_panel = monthly_panel[monthly_panel['country_code'].notna()].copy()
annual_panel = annual_panel[annual_panel['country_code'].notna()].copy()

monthly_panel['year'] = monthly_panel['date'].dt.year


# ============================================================
# Save outputs
# ============================================================

print("\n" + "=" * 60)
print("SAVING PROCESSED DATA")
print("=" * 60)

monthly_path = os.path.join(PROCESSED_DIR, "inflation_panel_monthly.csv")
monthly_panel.to_csv(monthly_path, index=False)
size_mb = os.path.getsize(monthly_path) / (1024*1024)
print(f"\n  Saved: {monthly_path}")
print(f"  Size: {size_mb:.1f} MB | Rows: {len(monthly_panel):,} | Countries: {monthly_panel['country_code'].nunique()}")

annual_path = os.path.join(PROCESSED_DIR, "inflation_panel_annual.csv")
annual_panel.to_csv(annual_path, index=False)
size_mb = os.path.getsize(annual_path) / (1024*1024)
print(f"\n  Saved: {annual_path}")
print(f"  Size: {size_mb:.1f} MB | Rows: {len(annual_panel):,} | Countries: {annual_panel['country_code'].nunique()}")


# ============================================================
# Summary
# ============================================================

print("\n" + "=" * 60)
print("STEP 2 COMPLETE")
print("=" * 60)

print(f"""
Monthly panel:  {monthly_panel['country_code'].nunique()} countries, {monthly_panel['date'].min().strftime('%Y-%m')} to {monthly_panel['date'].max().strftime('%Y-%m')}, {len(monthly_panel):,} rows
Annual panel:   {annual_panel['country_code'].nunique()} countries, {annual_panel['year'].min()} to {annual_panel['year'].max()}, {len(annual_panel):,} rows
""")
