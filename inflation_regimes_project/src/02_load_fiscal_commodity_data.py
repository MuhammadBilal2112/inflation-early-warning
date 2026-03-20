"""
==============================================================================
STEP 3: LOAD FISCAL SPACE AND COMMODITY DATA, THEN MERGE EVERYTHING
==============================================================================

  1. Loads all 23 fiscal space indicators + country classifications
  2. Loads 71 commodity price series and builds composite indices
  3. Merges everything with the inflation data from Step 2
  4. Produces two master panels ready for analysis

REQUIRES: Step 2 completed
OUTPUT:
  - data/processed/fiscal_panel.csv
  - data/processed/commodity_monthly.csv
  - data/processed/master_panel_monthly.csv
  - data/processed/master_panel_annual.csv

RUN: python3 Step03_Load_Fiscal_Commodity_Data.py
TIME: 1-2 minutes
==============================================================================
"""

import pandas as pd
import numpy as np
import openpyxl
import os
import warnings
warnings.filterwarnings('ignore')

BASE_DIR = "inflation_regimes_project"
RAW_DIR = os.path.join(BASE_DIR, "data", "raw")
PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")

for f in ["inflation_panel_monthly.csv", "inflation_panel_annual.csv"]:
    path = os.path.join(PROCESSED_DIR, f)
    if os.path.exists(path):
        print(f"  Found Step 2 output: {f}")
    else:
        print(f"  ERROR: Missing {f} — run Step 2 first!")


# ============================================================
# Load fiscal space indicators
# ============================================================

print("\n" + "=" * 60)
print("LOADING FISCAL SPACE DATA")
print("=" * 60)

fiscal_file = os.path.join(RAW_DIR, "Fiscalspacedata.xlsx")

fiscal_sheets = {
    'ggdy': 'debt_gdp',
    'pby': 'primary_balance',
    'cby': 'cycl_adj_balance',
    'fby': 'fiscal_balance',
    'dfggd': 'debt_to_tax',
    'dffb': 'fbal_to_tax',
    'ggdma': 'debt_to_ma_gdp',
    'fbma': 'fbal_to_ma_gdp',
    'fxsovsh': 'fx_debt_share',
    'secnres': 'nonres_securities',
    'fordebtsh': 'nonres_debt_share',
    'concggd': 'concessional_share',
    'avglife': 'avg_maturity',
    'debtduey': 'shortterm_debt_gdp',
    'xtdebty': 'ext_debt_gdp',
    'fxdebtall': 'fx_ext_debt_share',
    'prdebty': 'private_ext_debt',
    'pscy': 'private_credit_gdp',
    'stdebtall': 'st_ext_debt_share',
    'stdebtres': 'st_debt_reserves',
    'xtdebtres': 'ext_debt_reserves',
    'xtdebtrxg': 'ext_debt_res_exgold',
    'sovrate': 'sovereign_rating',
}

def load_fiscal_sheet(filepath, sheet_name, var_name):
    """Load one fiscal space sheet and reshape to long format."""

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[sheet_name]

    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    year_info = []
    for i, h in enumerate(header):
        if isinstance(h, str) and h.strip().isdigit():
            yr = int(h.strip())
            if 1980 <= yr <= 2030:
                year_info.append((i, yr))

    records = []
    country_info = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        country_code = row[0]
        country_name = row[2]

        if not country_code or not country_name:
            continue

        cc = str(country_code).strip()
        cn = str(country_name).strip()

        if cc not in country_info:
            country_info[cc] = {
                'country_name': cn,
                'country_group': str(row[3]).strip() if row[3] else None,
                'region': str(row[4]).strip() if row[4] else None,
                'income_group': str(row[5]).strip() if row[5] else None,
            }

        for col_idx, year in year_info:
            value = row[col_idx] if col_idx < len(row) else None
            if value is not None and isinstance(value, (int, float)):
                records.append({
                    'country_code': cc,
                    'year': year,
                    var_name: float(value)
                })

    wb.close()

    df = pd.DataFrame(records)
    return df, country_info


all_fiscal_dfs = {}
all_country_info = {}

for sheet_name, var_name in fiscal_sheets.items():
    print(f"  Loading {sheet_name:15s} -> {var_name}...", end="", flush=True)
    df, cinfo = load_fiscal_sheet(fiscal_file, sheet_name, var_name)
    all_fiscal_dfs[var_name] = df
    all_country_info.update(cinfo)

    n_c = df['country_code'].nunique() if not df.empty else 0
    print(f" {n_c} countries, {len(df):,} obs")


# ============================================================
# Merge fiscal variables
# ============================================================

print("\nMerging all fiscal variables...")

fiscal_panel = all_fiscal_dfs['debt_gdp'].copy()

for var_name, df in all_fiscal_dfs.items():
    if var_name == 'debt_gdp' or df.empty:
        continue
    fiscal_panel = fiscal_panel.merge(
        df[['country_code', 'year', var_name]],
        on=['country_code', 'year'],
        how='outer'
    )

classifications = pd.DataFrame.from_dict(all_country_info, orient='index')
classifications.index.name = 'country_code'
classifications = classifications.reset_index()

fiscal_panel = fiscal_panel.merge(
    classifications,
    on='country_code',
    how='left'
)

fiscal_panel = fiscal_panel.sort_values(['country_code', 'year']).reset_index(drop=True)

print(f"\nFiscal panel created:")
print(f"  Shape: {fiscal_panel.shape}")
print(f"  Countries: {fiscal_panel['country_code'].nunique()}")
print(f"  Years: {fiscal_panel['year'].min()} to {fiscal_panel['year'].max()}")
print(f"\n  Coverage by variable (% non-missing):")

good_coverage = []
moderate_coverage = []
sparse_coverage = []

fiscal_var_names = list(fiscal_sheets.values())
for var in fiscal_var_names:
    if var in fiscal_panel.columns:
        pct = fiscal_panel[var].notna().mean() * 100
        label = f"    {var:25s}: {pct:5.1f}%"
        if pct >= 50:
            good_coverage.append((var, pct, label))
        elif pct >= 20:
            moderate_coverage.append((var, pct, label))
        else:
            sparse_coverage.append((var, pct, label))

print(f"\n  GOOD (>50%):")
for _, _, label in good_coverage:
    print(label)

print(f"\n  MODERATE (20-50%):")
for _, _, label in moderate_coverage:
    print(label)

print(f"\n  SPARSE (<20%):")
for _, _, label in sparse_coverage:
    print(label)


# ============================================================
# Load commodity price data
# ============================================================

print("\n" + "=" * 60)
print("LOADING COMMODITY PRICE DATA")
print("=" * 60)

commodity_file = os.path.join(RAW_DIR, "CMOHistoricalDataMonthly.xlsx")

print("  Reading Monthly Prices sheet...", end="", flush=True)
wb = openpyxl.load_workbook(commodity_file, read_only=True)
ws = wb['Monthly Prices']

header_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
unit_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]

commodity_cols = {}
for i, (name, unit) in enumerate(zip(header_row, unit_row)):
    if i == 0:
        continue
    if name and str(name).strip():
        commodity_cols[i] = (str(name).strip(), str(unit).strip() if unit else '')

records = []
for row in ws.iter_rows(min_row=7, values_only=True):
    date_str = row[0]
    if not date_str or not isinstance(date_str, str) or 'M' not in date_str:
        continue

    try:
        year = int(date_str.split('M')[0])
        month = int(date_str.split('M')[1])
    except (ValueError, IndexError):
        continue

    record = {'date_str': date_str, 'year': year, 'month': month}

    for col_idx, (name, unit) in commodity_cols.items():
        value = row[col_idx] if col_idx < len(row) else None
        if value is not None and isinstance(value, (int, float)):
            record[name] = float(value)

    records.append(record)

wb.close()

commodity_raw = pd.DataFrame(records)
commodity_raw['date'] = pd.to_datetime(
    commodity_raw['year'].astype(str) + '-' +
    commodity_raw['month'].astype(str).str.zfill(2) + '-01'
)

print(f" {len(commodity_raw)} months, {len(commodity_cols)} series")
print(f"  Date range: {commodity_raw['date'].min().strftime('%Y-%m')} to {commodity_raw['date'].max().strftime('%Y-%m')}")


# ============================================================
# Build composite commodity indices
# ============================================================

print("\nBuilding composite commodity indices...")

def build_index(df, columns, index_name):
    """
    Build a simple average index from available columns.
    Normalises each series to 100 at the 2010 average.
    """
    available = [c for c in columns if c in df.columns]
    if not available:
        print(f"  WARNING: No columns found for {index_name}")
        return pd.Series(dtype=float)

    normalised = pd.DataFrame()
    for col in available:
        series = df.set_index('date')[col]
        base_value = series['2010-01-01':'2010-12-01'].mean()
        if base_value and base_value > 0:
            normalised[col] = (series / base_value) * 100

    if normalised.empty:
        return pd.Series(dtype=float)

    index = normalised.mean(axis=1)
    print(f"  {index_name}: built from {len(available)} series ({', '.join(available[:3])}{'...' if len(available)>3 else ''})")
    return index


energy_cols = ['Crude oil, average', 'Coal, Australian', 'Natural gas, Europe']
food_cols = ['Wheat, US HRW', 'Maize', 'Rice, Thai 5%', 'Soybeans',
             'Palm oil', 'Sugar, world', 'Chicken **', 'Beef **']
metals_cols = ['Aluminum', 'Copper', 'Iron ore, cfr spot', 'Zinc', 'Nickel', 'Lead']
fertiliser_cols = ['DAP', 'Urea', 'Potassium chloride **', 'Phosphate rock', 'TSP']

commodity_features = pd.DataFrame({'date': commodity_raw['date']})

for cols, name in [(energy_cols, 'energy_index'),
                    (food_cols, 'food_commodity_index'),
                    (metals_cols, 'metals_index'),
                    (fertiliser_cols, 'fertiliser_index')]:
    idx = build_index(commodity_raw, cols, name)
    if not idx.empty:
        commodity_features[name] = idx.values

if 'Crude oil, average' in commodity_raw.columns:
    commodity_features['oil_price'] = commodity_raw['Crude oil, average'].values

if 'Wheat, US HRW' in commodity_raw.columns:
    commodity_features['wheat_price'] = commodity_raw['Wheat, US HRW'].values
if 'Rice, Thai 5%' in commodity_raw.columns:
    commodity_features['rice_price'] = commodity_raw['Rice, Thai 5%'].values


# ============================================================
# Compute commodity price change features
# ============================================================

print("\nComputing commodity price change features...")

commodity_features = commodity_features.sort_values('date').reset_index(drop=True)

index_cols = ['energy_index', 'food_commodity_index', 'metals_index',
              'fertiliser_index', 'oil_price']

for col in index_cols:
    if col not in commodity_features.columns:
        continue

    commodity_features[f'{col}_3m_chg'] = commodity_features[col].pct_change(3) * 100
    commodity_features[f'{col}_12m_chg'] = commodity_features[col].pct_change(12) * 100

    monthly_chg = commodity_features[col].pct_change() * 100
    commodity_features[f'{col}_6m_vol'] = monthly_chg.rolling(6).std()

    commodity_features[f'{col}_log'] = np.log(commodity_features[col].clip(lower=0.01))

print(f"  Commodity features: {commodity_features.shape[1]} columns total")


# ============================================================
# Build master monthly panel
# ============================================================

print("\n" + "=" * 60)
print("BUILDING MASTER MONTHLY PANEL")
print("=" * 60)

inflation_monthly = pd.read_csv(
    os.path.join(PROCESSED_DIR, "inflation_panel_monthly.csv"),
    parse_dates=['date']
)
print(f"  Loaded inflation monthly: {inflation_monthly.shape}")

# Commodity data is global — merge on date only
master_monthly = inflation_monthly.merge(
    commodity_features,
    on='date',
    how='left'
)
print(f"  After commodity merge: {master_monthly.shape}")

# Fiscal data is annual — assign by year
fiscal_for_merge = fiscal_panel.copy()

core_fiscal_vars = ['country_code', 'year', 'country_group', 'region', 'income_group',
                     'debt_gdp', 'primary_balance', 'fiscal_balance',
                     'ext_debt_gdp', 'st_debt_reserves', 'private_credit_gdp',
                     'sovereign_rating', 'concessional_share',
                     'fx_debt_share', 'nonres_debt_share', 'shortterm_debt_gdp']

core_fiscal_vars = [c for c in core_fiscal_vars if c in fiscal_for_merge.columns]
fiscal_for_merge = fiscal_for_merge[core_fiscal_vars].copy()

master_monthly = master_monthly.merge(
    fiscal_for_merge,
    on=['country_code', 'year'],
    how='left'
)
print(f"  After fiscal merge: {master_monthly.shape}")

master_monthly = master_monthly.sort_values(['country_code', 'date']).reset_index(drop=True)

n_total = len(master_monthly)
n_with_fiscal = master_monthly['debt_gdp'].notna().sum()
n_with_commodity = master_monthly['oil_price'].notna().sum()
n_with_classification = master_monthly['country_group'].notna().sum()

print(f"\n  Merge diagnostics:")
print(f"    Total rows: {n_total:,}")
print(f"    With fiscal data: {n_with_fiscal:,} ({n_with_fiscal/n_total*100:.1f}%)")
print(f"    With commodity data: {n_with_commodity:,} ({n_with_commodity/n_total*100:.1f}%)")
print(f"    With country classification: {n_with_classification:,} ({n_with_classification/n_total*100:.1f}%)")

if 'country_group' in master_monthly.columns:
    print(f"\n  Country group breakdown:")
    group_counts = master_monthly.groupby('country_group')['country_code'].nunique()
    for group, count in group_counts.items():
        print(f"    {group}: {count} countries")


# ============================================================
# Build master annual panel
# ============================================================

print("\n" + "=" * 60)
print("BUILDING MASTER ANNUAL PANEL")
print("=" * 60)

inflation_annual = pd.read_csv(
    os.path.join(PROCESSED_DIR, "inflation_panel_annual.csv")
)
print(f"  Loaded inflation annual: {inflation_annual.shape}")

commodity_annual = commodity_features.copy()
commodity_annual['year'] = commodity_annual['date'].dt.year
commodity_annual = commodity_annual.drop(columns=['date']).groupby('year').mean().reset_index()
print(f"  Computed annual commodity averages: {commodity_annual.shape}")

master_annual = inflation_annual.merge(
    fiscal_panel,
    on=['country_code', 'year'],
    how='outer',
    suffixes=('', '_fiscal')
)

if 'country_name_fiscal' in master_annual.columns:
    master_annual['country_name'] = master_annual['country_name'].fillna(master_annual['country_name_fiscal'])
    master_annual = master_annual.drop(columns=['country_name_fiscal'])

master_annual = master_annual.merge(
    commodity_annual,
    on='year',
    how='left'
)

master_annual = master_annual.sort_values(['country_code', 'year']).reset_index(drop=True)
print(f"  Master annual panel: {master_annual.shape}")
print(f"  Countries: {master_annual['country_code'].nunique()}")
print(f"  Years: {master_annual['year'].min()} to {master_annual['year'].max()}")


# ============================================================
# Save outputs
# ============================================================

print("\n" + "=" * 60)
print("SAVING ALL PROCESSED DATA")
print("=" * 60)

fiscal_path = os.path.join(PROCESSED_DIR, "fiscal_panel.csv")
fiscal_panel.to_csv(fiscal_path, index=False)
print(f"  Saved: fiscal_panel.csv ({len(fiscal_panel):,} rows, {fiscal_panel['country_code'].nunique()} countries)")

commodity_path = os.path.join(PROCESSED_DIR, "commodity_monthly.csv")
commodity_features.to_csv(commodity_path, index=False)
print(f"  Saved: commodity_monthly.csv ({len(commodity_features):,} rows, {commodity_features.shape[1]} columns)")

master_monthly_path = os.path.join(PROCESSED_DIR, "master_panel_monthly.csv")
master_monthly.to_csv(master_monthly_path, index=False)
size_mb = os.path.getsize(master_monthly_path) / (1024*1024)
print(f"  Saved: master_panel_monthly.csv ({len(master_monthly):,} rows, {size_mb:.1f} MB)")

master_annual_path = os.path.join(PROCESSED_DIR, "master_panel_annual.csv")
master_annual.to_csv(master_annual_path, index=False)
size_mb = os.path.getsize(master_annual_path) / (1024*1024)
print(f"  Saved: master_panel_annual.csv ({len(master_annual):,} rows, {size_mb:.1f} MB)")


print(f"\n{'='*60}")
print("STEP 3 COMPLETE")
print(f"{'='*60}")
print(f"""
  master_panel_monthly.csv  — {len(master_monthly):,} rows, {master_monthly.shape[1]} columns
  master_panel_annual.csv   — {len(master_annual):,} rows, {master_annual.shape[1]} columns
""")
