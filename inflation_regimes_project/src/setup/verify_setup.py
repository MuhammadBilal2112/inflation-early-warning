"""
STEP 1C: Verify Everything Is Working
======================================
Run this LAST to confirm your setup is complete.

HOW TO RUN:
  python verify_setup.py

This script checks:
  1. All libraries are installed and importable
  2. Your data files are in the right place
  3. The data files can be read correctly
  4. A quick sanity check on the data contents

If everything passes, you are ready for Step 2!
"""

import os
import sys

def print_header(text):
    print()
    print("=" * 60)
    print(f"  {text}")
    print("=" * 60)

def print_pass(text):
    print(f"  [PASS] {text}")

def print_fail(text):
    print(f"  [FAIL] {text}")

def print_warn(text):
    print(f"  [WARN] {text}")

def print_info(text):
    print(f"  [INFO] {text}")

def verify_setup():
    """Run all verification checks."""
    
    all_passed = True
    
    # ============================================================
    # CHECK 1: Python version
    # ============================================================
    print_header("CHECK 1: Python Version")
    
    major, minor = sys.version_info[:2]
    if major == 3 and minor >= 10:
        print_pass(f"Python {major}.{minor} (3.10+ required)")
    elif major == 3 and minor >= 8:
        print_warn(f"Python {major}.{minor} (works but 3.10+ recommended)")
    else:
        print_fail(f"Python {major}.{minor} (need 3.10+, some features may not work)")
        all_passed = False
    
    # ============================================================
    # CHECK 2: All libraries
    # ============================================================
    print_header("CHECK 2: Required Libraries")
    
    libraries = {
        'numpy': 'Core math operations',
        'pandas': 'Data manipulation',
        'openpyxl': 'Excel file reading',
        'matplotlib': 'Static charts',
        'seaborn': 'Statistical plots',
        'scipy': 'Scientific computing',
        'sklearn': 'Machine learning core',
        'xgboost': 'Gradient boosted trees',
        'lightgbm': 'Fast gradient boosting',
        'shap': 'Model interpretability',
        'hmmlearn': 'Hidden Markov Models',
        'tslearn': 'Time series clustering',
        'plotly': 'Interactive charts',
        'statsmodels': 'Statistical tests',
    }
    
    # Optional libraries (nice to have)
    optional = {
        'geopandas': 'Geographic maps',
        'jupyter': 'Jupyter notebooks',
    }
    
    missing_required = []
    for import_name, description in libraries.items():
        try:
            mod = __import__(import_name)
            ver = getattr(mod, '__version__', '?')
            print_pass(f"{import_name:15s} v{ver:10s}  ({description})")
        except ImportError:
            print_fail(f"{import_name:15s} NOT FOUND   ({description})")
            missing_required.append(import_name)
            all_passed = False
    
    print()
    print("  Optional libraries:")
    for import_name, description in optional.items():
        try:
            mod = __import__(import_name)
            ver = getattr(mod, '__version__', '?')
            print_pass(f"{import_name:15s} v{ver:10s}  ({description})")
        except ImportError:
            print_warn(f"{import_name:15s} NOT FOUND   ({description}) - not critical")
    
    if missing_required:
        print()
        print(f"  Install missing libraries with:")
        pip_names = {'sklearn': 'scikit-learn'}  # special case
        install_names = [pip_names.get(n, n) for n in missing_required]
        print(f"  pip install {' '.join(install_names)}")
    
    # ============================================================
    # CHECK 3: Data files
    # ============================================================
    print_header("CHECK 3: Data Files")
    
    # Try multiple possible locations
    possible_data_dirs = [
        "inflation_regimes_project/data/raw",
        "data/raw",
        ".",
    ]
    
    required_files = {
        'Inflationdata.xlsx': 'Inflation Database (209 countries, 6 measures)',
        'Fiscalspacedata.xlsx': 'Fiscal Space Database (204 countries, 23 indicators)',
        'CMOHistoricalDataMonthly.xlsx': 'Commodity Prices Monthly (71 series)',
        'CMOHistoricalDataAnnual.xlsx': 'Commodity Prices Annual',
    }
    
    data_dir = None
    for candidate in possible_data_dirs:
        if os.path.isdir(candidate):
            files_found = sum(1 for f in required_files if os.path.exists(os.path.join(candidate, f)))
            if files_found > 0:
                data_dir = candidate
                break
    
    if data_dir is None:
        print_fail("Could not find data directory!")
        print_info("Expected data files in: inflation_regimes_project/data/raw/")
        print_info("Make sure you copied your Excel files there.")
        all_passed = False
    else:
        print_info(f"Looking in: {os.path.abspath(data_dir)}/")
        print()
        for filename, description in required_files.items():
            filepath = os.path.join(data_dir, filename)
            if os.path.exists(filepath):
                size_mb = os.path.getsize(filepath) / (1024 * 1024)
                print_pass(f"{filename:40s} ({size_mb:.1f} MB) - {description}")
            else:
                print_fail(f"{filename:40s} NOT FOUND")
                all_passed = False
    
    # ============================================================
    # CHECK 4: Quick data read test
    # ============================================================
    print_header("CHECK 4: Data Readability Test")
    
    if data_dir and 'pandas' not in missing_required and 'openpyxl' not in missing_required:
        import pandas as pd
        import openpyxl
        
        # Test 1: Read inflation data
        infl_path = os.path.join(data_dir, 'Inflationdata.xlsx')
        if os.path.exists(infl_path):
            try:
                wb = openpyxl.load_workbook(infl_path, read_only=True)
                sheets = wb.sheetnames
                print_pass(f"Inflation data readable: {len(sheets)} sheets")
                
                # Check key sheets exist
                expected_sheets = ['hcpi_m', 'hcpi_a', 'fcpi_m', 'fcpi_a', 'ecpi_m', 'ecpi_a', 
                                   'ccpi_m', 'ccpi_a', 'ppi_m', 'ppi_a', 'def_a', 'Aggregate']
                for s in expected_sheets:
                    if s in sheets:
                        print_pass(f"  Sheet '{s}' found")
                    else:
                        print_fail(f"  Sheet '{s}' MISSING")
                        all_passed = False
                
                # Quick content check - count UK data
                ws = wb['hcpi_a']
                uk_found = False
                for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
                    if row[0] == 'GBR':
                        uk_found = True
                        print_pass(f"  UK data found in hcpi_a: {row[2]}, Type: {row[3]}")
                        break
                if not uk_found:
                    print_warn("  Could not find UK data (GBR) in hcpi_a")
                
                # Verify monthly data is INDEX
                ws_m = wb['hcpi_m']
                first_row = list(ws_m.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                indicator_type = first_row[3]
                print_pass(f"  Monthly hcpi_m indicator type: '{indicator_type}' (should be 'Index')")
                if indicator_type != 'Index':
                    print_warn("  Expected 'Index' - check column structure!")
                
                wb.close()
                
            except Exception as e:
                print_fail(f"Error reading inflation data: {e}")
                all_passed = False
        
        # Test 2: Read fiscal data
        fiscal_path = os.path.join(data_dir, 'Fiscalspacedata.xlsx')
        if os.path.exists(fiscal_path):
            try:
                wb2 = openpyxl.load_workbook(fiscal_path, read_only=True)
                print_pass(f"Fiscal data readable: {len(wb2.sheetnames)} sheets")
                
                # Check country classifications exist
                ws_f = wb2['ggdy']
                first_row = list(ws_f.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                print_pass(f"  Country group column present: '{first_row[3]}'")
                print_pass(f"  Region column present: '{first_row[4]}'")
                print_pass(f"  Income group column present: '{first_row[5]}'")
                wb2.close()
                
            except Exception as e:
                print_fail(f"Error reading fiscal data: {e}")
                all_passed = False
        
        # Test 3: Read commodity data
        cmo_path = os.path.join(data_dir, 'CMOHistoricalDataMonthly.xlsx')
        if os.path.exists(cmo_path):
            try:
                wb3 = openpyxl.load_workbook(cmo_path, read_only=True)
                print_pass(f"Commodity data readable: {len(wb3.sheetnames)} sheets")
                
                ws_c = wb3['Monthly Prices']
                header = list(ws_c.iter_rows(min_row=5, max_row=5, max_col=5, values_only=True))[0]
                print_pass(f"  First commodity: '{header[1]}'")
                wb3.close()
                
            except Exception as e:
                print_fail(f"Error reading commodity data: {e}")
                all_passed = False
    else:
        print_warn("Skipping data read test (missing pandas/openpyxl or data directory)")
    
    # ============================================================
    # CHECK 5: Quick ML test
    # ============================================================
    print_header("CHECK 5: Quick ML Functionality Test")
    
    if 'sklearn' not in missing_required:
        try:
            from sklearn.ensemble import RandomForestClassifier
            from sklearn.datasets import make_classification
            import numpy as np
            
            # Generate tiny fake dataset and train
            X, y = make_classification(n_samples=100, n_features=5, random_state=42)
            rf = RandomForestClassifier(n_estimators=10, random_state=42)
            rf.fit(X, y)
            accuracy = rf.score(X, y)
            print_pass(f"scikit-learn RandomForest works (accuracy: {accuracy:.2f})")
        except Exception as e:
            print_fail(f"scikit-learn test failed: {e}")
            all_passed = False
    
    if 'xgboost' not in missing_required:
        try:
            import xgboost as xgb
            import numpy as np
            
            X = np.random.randn(100, 5)
            y = (X[:, 0] > 0).astype(int)
            model = xgb.XGBClassifier(n_estimators=10, verbosity=0, use_label_encoder=False)
            model.fit(X, y)
            print_pass(f"XGBoost works (v{xgb.__version__})")
        except Exception as e:
            print_fail(f"XGBoost test failed: {e}")
            all_passed = False
    
    if 'hmmlearn' not in missing_required:
        try:
            from hmmlearn import hmm
            import numpy as np
            
            model = hmm.GaussianHMM(n_components=3, n_iter=10)
            X = np.random.randn(100, 2)
            model.fit(X)
            print_pass(f"hmmlearn works (Hidden Markov Models ready)")
        except Exception as e:
            print_fail(f"hmmlearn test failed: {e}")
            all_passed = False
    
    if 'shap' not in missing_required:
        try:
            import shap
            print_pass(f"SHAP importable (v{shap.__version__})")
        except Exception as e:
            print_fail(f"SHAP test failed: {e}")
            all_passed = False
    
    # ============================================================
    # FINAL VERDICT
    # ============================================================
    print_header("FINAL RESULT")
    
    if all_passed:
        print()
        print("  ############################################")
        print("  #                                          #")
        print("  #   ALL CHECKS PASSED - YOU ARE READY!     #")
        print("  #                                          #")
        print("  #   Proceed to Step 2: Load Inflation Data #")
        print("  #                                          #")
        print("  ############################################")
        print()
        print("  To start Step 2:")
        print("  1. Open a terminal in your project folder")
        print("  2. Run: jupyter notebook")
        print("  3. Create a new notebook called 'Step02_Load_Inflation_Data.ipynb'")
        print("  4. Come back to Claude and say 'ready for step 2'")
    else:
        print()
        print("  SOME CHECKS FAILED - see details above.")
        print("  Fix the issues and run this script again.")
        print("  If you get stuck, copy the full output and")
        print("  share it with Claude for troubleshooting.")
    print()

if __name__ == "__main__":
    verify_setup()
